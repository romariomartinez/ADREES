from __future__ import annotations

import argparse
import base64
import hashlib
import json
import mimetypes
import os
import re
import secrets
import sqlite3
import importlib
from copy import deepcopy
from datetime import datetime
from http import HTTPStatus
from http.cookies import SimpleCookie
from email.message import Message
from io import BytesIO
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer

try:
    import psycopg
    from psycopg.rows import dict_row
except ImportError:
    psycopg = None
    dict_row = None

try:
    import openpyxl
except ImportError:
    openpyxl = None


ROOT = Path(__file__).resolve().parent
DATABASE_URL = (
    os.environ.get("DATABASE_URL")
    or os.environ.get("SUPABASE_DATABASE_URL")
    or os.environ.get("POSTGRES_URL")
    or ""
)
USE_POSTGRES = bool(DATABASE_URL)
IS_VERCEL = bool(os.environ.get("VERCEL") or os.environ.get("VERCEL_ENV"))
DATA_DIR = Path(
    os.environ.get("ADRES_DATA_DIR")
    or os.environ.get("RAILWAY_VOLUME_MOUNT_PATH")
    or ROOT / "data"
)
SCHEMA_PATH = ROOT / "data" / "schema.json"
DIVIPOLA_PATH = ROOT / "data" / "divipola.json"
STATIC_DIR = ROOT / "static"
TEMPLATE_DIR = ROOT / "templates"
EXPORT_DIR = Path(os.environ.get("ADRES_EXPORT_DIR", DATA_DIR / "exports"))
DB_PATH = Path(os.environ.get("ADRES_DB_PATH", DATA_DIR / "app.db"))
SAVE_EXPORT_COPY = os.environ.get("ADRES_SAVE_EXPORT_COPY")
if SAVE_EXPORT_COPY is None:
    SAVE_EXPORT_COPY_ENABLED = not IS_VERCEL
else:
    SAVE_EXPORT_COPY_ENABLED = SAVE_EXPORT_COPY.strip().lower() not in {"0", "false", "no", "off"}
DB_INTEGRITY_ERRORS = (sqlite3.IntegrityError,)
if psycopg is not None:
    DB_INTEGRITY_ERRORS = (sqlite3.IntegrityError, psycopg.IntegrityError)
DB_READY = False
DB_BOOTSTRAPPING = False
DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
TIME_RE = re.compile(r"^(?:[01]\d|2[0-3]):[0-5]\d$")
DIGITS_RE = re.compile(r"^\d+$")
INVOICE_PREFIX = "FVEE"
ROLE_SUPER_ADMIN = "super_admin"
ROLE_FACTURADOR = "facturador"
VALID_ROLES = {ROLE_SUPER_ADMIN, ROLE_FACTURADOR}
SESSION_COOKIE = "adres_session"
DIVIPOLA_FIELD_NAMES = {
    "Codigo_municipio_residencia_victima",
    "Codigo_municipio_ocurrencia_evento",
    "Codigo_del_municipio_de_residencia_del_propietario",
    "Codigo_del_municipio_de_residencia_del_conductor",
}
FREQUENT_FIELD_NAMES = {
    "NIT_PRESTADOR",
    "Codigo_municipio_residencia_victima",
    "Codigo_municipio_ocurrencia_evento",
    "Codigo_del_municipio_de_residencia_del_propietario",
    "Codigo_del_municipio_de_residencia_del_conductor",
    "Codigo_de_la_aseguradora",
    "Codigo_de_habilitacion_del_prestador_que_remite",
    "Codigo_de_habilitacion_del_prestador_que_recibe",
    "Codigo_de_habilitacion_del_prestador_que_recibe_transporte_primario",
    "Direccion_residencia_victima",
    "Direccion_de_ocurrencia_evento",
    "Direccion_de_residencia_del_propietario",
    "Direccion_de_residencia_del_conductor",
}


def load_schema() -> dict:
    with SCHEMA_PATH.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def load_divipola() -> list[dict]:
    if not DIVIPOLA_PATH.exists():
        return []
    with DIVIPOLA_PATH.open("r", encoding="utf-8") as fh:
        items = json.load(fh)
    return sorted(items, key=lambda item: item["code"])


SCHEMA = load_schema()
DIVIPOLA_ITEMS = load_divipola()
DIVIPOLA_CODES = {item["code"] for item in DIVIPOLA_ITEMS}


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def safe_error_text(exc: Exception) -> str:
    text = str(exc) or exc.__class__.__name__
    if DATABASE_URL:
        text = text.replace(DATABASE_URL, "[DATABASE_URL]")
    return re.sub(r"(postgres(?:ql)?://[^:\s]+:)[^@\s]+@", r"\1***@", text)


def get_openpyxl():
    global openpyxl
    if openpyxl is None:
        try:
            openpyxl = importlib.import_module("openpyxl")
        except ImportError as exc:
            raise RuntimeError(
                "Falta la dependencia openpyxl. Revisa que Vercel instale requirements.txt."
            ) from exc
    return openpyxl


def validate_database_url():
    if IS_VERCEL and re.search(r"@db\.[^/:]+\.supabase\.co:5432\b", DATABASE_URL):
        raise RuntimeError(
            "En Vercel no uses la conexion directa de Supabase db.[proyecto].supabase.co:5432. "
            "Usa la cadena Transaction pooler de Supabase, normalmente aws-[region].pooler.supabase.com:6543."
        )
    if IS_VERCEL and "pooler.supabase.com:6543" in DATABASE_URL and re.search(r"^postgresql://postgres:", DATABASE_URL):
        raise RuntimeError(
            "La URL del Transaction pooler de Supabase debe usar usuario postgres.PROJECT_REF, no solo postgres. "
            "En este proyecto debe empezar parecido a postgresql://postgres.ixrdhxqoqkdoayzrahtb:TU_CLAVE@aws-...pooler.supabase.com:6543/postgres?sslmode=require."
        )


def db_sql(sql: str) -> str:
    if USE_POSTGRES:
        return sql.replace("?", "%s")
    return sql


def like_op() -> str:
    return "ILIKE" if USE_POSTGRES else "LIKE"


class DatabaseConnection:
    def __init__(self, conn):
        self.conn = conn

    def __enter__(self):
        self.conn.__enter__()
        return self

    def __exit__(self, exc_type, exc, traceback):
        try:
            return self.conn.__exit__(exc_type, exc, traceback)
        finally:
            self.conn.close()

    def execute(self, sql: str, params: tuple | list = ()):
        return self.conn.execute(db_sql(sql), params)

    def insert_and_get_id(self, sql: str, params: tuple | list = ()) -> int:
        if USE_POSTGRES:
            cursor = self.execute(f"{sql.strip()} RETURNING id", params)
            return int(cursor.fetchone()["id"])
        cursor = self.execute(sql, params)
        return int(cursor.lastrowid)


def db_connect() -> DatabaseConnection:
    if not DB_BOOTSTRAPPING:
        ensure_db()

    if USE_POSTGRES:
        validate_database_url()
        if psycopg is None:
            raise RuntimeError(
                "Falta la dependencia psycopg para conectar con Supabase/Postgres. "
                "Instala con: python -m pip install -r requirements.txt"
            )
        conn = psycopg.connect(DATABASE_URL, row_factory=dict_row, prepare_threshold=None)
        return DatabaseConnection(conn)

    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    return DatabaseConnection(conn)


def ensure_db():
    global DB_READY, DB_BOOTSTRAPPING
    if DB_READY or DB_BOOTSTRAPPING:
        return
    DB_BOOTSTRAPPING = True
    try:
        init_db()
        DB_READY = True
    finally:
        DB_BOOTSTRAPPING = False


def table_columns(conn: DatabaseConnection, table_name: str) -> set[str]:
    if USE_POSTGRES:
        rows = conn.execute(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = ?
            """,
            (table_name,),
        ).fetchall()
        return {row["column_name"] for row in rows}
    rows = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    return {row["name"] for row in rows}


def init_db():
    with db_connect() as conn:
        id_type = "BIGSERIAL" if USE_POSTGRES else "INTEGER"
        user_id_type = "BIGINT" if USE_POSTGRES else "INTEGER"
        primary_key = "PRIMARY KEY" if USE_POSTGRES else "PRIMARY KEY AUTOINCREMENT"
        conn.execute(
            f"""
            CREATE TABLE IF NOT EXISTS users (
                id {id_type} {primary_key},
                username TEXT NOT NULL UNIQUE,
                display_name TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'facturador',
                active INTEGER NOT NULL DEFAULT 1,
                password_salt TEXT NOT NULL,
                password_hash TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
            """
        )
        columns = table_columns(conn, "users")
        if "role" not in columns:
            conn.execute("ALTER TABLE users ADD COLUMN role TEXT NOT NULL DEFAULT 'facturador'")
        if "active" not in columns:
            conn.execute("ALTER TABLE users ADD COLUMN active INTEGER NOT NULL DEFAULT 1")
        first_user = conn.execute("SELECT id FROM users ORDER BY id LIMIT 1").fetchone()
        super_admin = conn.execute(
            "SELECT id FROM users WHERE role = ? LIMIT 1",
            (ROLE_SUPER_ADMIN,),
        ).fetchone()
        if first_user and not super_admin:
            conn.execute("UPDATE users SET role = ? WHERE id = ?", (ROLE_SUPER_ADMIN, first_user["id"]))
        conn.execute(
            f"""
            CREATE TABLE IF NOT EXISTS sessions (
                id TEXT PRIMARY KEY,
                user_id {user_id_type} NOT NULL,
                created_at TEXT NOT NULL,
                last_seen_at TEXT NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_sessions_user
            ON sessions(user_id)
            """
        )
        conn.execute(
            f"""
            CREATE TABLE IF NOT EXISTS invoice_records (
                id {id_type} {primary_key},
                user_id {user_id_type} NOT NULL,
                template_id TEXT NOT NULL,
                invoice_number TEXT NOT NULL,
                filename TEXT NOT NULL,
                row_count INTEGER NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
            """
        )
        conn.execute(
            f"""
            CREATE TABLE IF NOT EXISTS drafts (
                id {id_type} {primary_key},
                user_id {user_id_type} NOT NULL,
                template_id TEXT NOT NULL,
                invoice_number TEXT NOT NULL,
                payload_json TEXT NOT NULL,
                row_count INTEGER NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_invoice_records_template_invoice
            ON invoice_records(template_id, invoice_number)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_drafts_user_template_invoice
            ON drafts(user_id, template_id, invoice_number)
            """
        )
        conn.execute(
            f"""
            CREATE TABLE IF NOT EXISTS frequent_values (
                id {id_type} {primary_key},
                user_id {user_id_type} NOT NULL,
                field_name TEXT NOT NULL,
                value TEXT NOT NULL,
                label TEXT NOT NULL,
                use_count INTEGER NOT NULL DEFAULT 1,
                last_used_at TEXT NOT NULL,
                UNIQUE(user_id, field_name, value),
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
            """
        )

def hash_password(password: str, salt: bytes | None = None) -> tuple[str, str]:
    salt = salt or secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 120_000)
    return (
        base64.b64encode(salt).decode("ascii"),
        base64.b64encode(digest).decode("ascii"),
    )


def verify_password(password: str, salt_text: str, hash_text: str) -> bool:
    salt = base64.b64decode(salt_text.encode("ascii"))
    _, candidate = hash_password(password, salt)
    return secrets.compare_digest(candidate, hash_text)


def public_user(row: dict | sqlite3.Row) -> dict:
    return {
        "id": row["id"],
        "username": row["username"],
        "displayName": row["display_name"],
        "role": row["role"],
        "active": bool(row["active"]),
        "isSuperAdmin": row["role"] == ROLE_SUPER_ADMIN,
    }


def user_count() -> int:
    with db_connect() as conn:
        row = conn.execute("SELECT COUNT(*) AS total FROM users").fetchone()
    return int(row["total"])


def create_user(username: str, password: str, display_name: str, role: str = ROLE_FACTURADOR) -> tuple[dict | None, str | None]:
    username = username.strip().lower()
    display_name = display_name.strip() or username
    role = role if role in VALID_ROLES else ROLE_FACTURADOR
    if not re.match(r"^[a-z0-9._-]{3,30}$", username):
        return None, "El usuario debe tener 3 a 30 caracteres: letras, numeros, punto, guion o guion bajo."
    if len(password) < 4:
        return None, "La clave debe tener minimo 4 caracteres."
    salt, password_hash = hash_password(password)
    try:
        with db_connect() as conn:
            user_id = conn.insert_and_get_id(
                """
                INSERT INTO users (username, display_name, role, active, password_salt, password_hash, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (username, display_name, role, 1, salt, password_hash, now_iso()),
            )
            row = conn.execute(
                "SELECT id, username, display_name, role, active FROM users WHERE id = ?",
                (user_id,),
            ).fetchone()
            return public_user(row), None
    except DB_INTEGRITY_ERRORS:
        return None, "Ese usuario ya existe."


def authenticate(username: str, password: str) -> dict | None:
    with db_connect() as conn:
        row = conn.execute(
            "SELECT * FROM users WHERE username = ?",
            (username.strip().lower(),),
        ).fetchone()
    if not row or not row["active"] or not verify_password(password, row["password_salt"], row["password_hash"]):
        return None
    return public_user(row)


def create_session(user: dict) -> str:
    session_id = secrets.token_urlsafe(32)
    timestamp = now_iso()
    with db_connect() as conn:
        conn.execute(
            """
            INSERT INTO sessions (id, user_id, created_at, last_seen_at)
            VALUES (?, ?, ?, ?)
            """,
            (session_id, user["id"], timestamp, timestamp),
        )
    return session_id


def delete_session(session_id: str):
    if not session_id:
        return
    with db_connect() as conn:
        conn.execute("DELETE FROM sessions WHERE id = ?", (session_id,))


def get_user_by_session(session_id: str) -> dict | None:
    if not session_id:
        return None
    with db_connect() as conn:
        row = conn.execute(
            """
            SELECT u.id, u.username, u.display_name, u.role, u.active
            FROM sessions s
            JOIN users u ON u.id = s.user_id
            WHERE s.id = ?
            """,
            (session_id,),
        ).fetchone()
        if not row:
            return None
        if not row["active"]:
            conn.execute("DELETE FROM sessions WHERE id = ?", (session_id,))
            return None
        conn.execute(
            "UPDATE sessions SET last_seen_at = ? WHERE id = ?",
            (now_iso(), session_id),
        )
    return public_user(row)


def record_export(user: dict, template_id: str, invoice_number: str, filename: str, row_count: int):
    with db_connect() as conn:
        conn.execute(
            """
            INSERT INTO invoice_records (user_id, template_id, invoice_number, filename, row_count, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (user["id"], template_id, invoice_number, filename, row_count, now_iso()),
        )


def find_duplicate_export(template_id: str, invoice_number: str) -> dict | None:
    if not invoice_number:
        return None
    with db_connect() as conn:
        row = conn.execute(
            """
            SELECT r.id, r.template_id, r.invoice_number, r.filename, r.row_count, r.created_at,
                   u.username, u.display_name
            FROM invoice_records r
            JOIN users u ON u.id = r.user_id
            WHERE r.template_id = ? AND r.invoice_number = ?
            ORDER BY r.id DESC
            LIMIT 1
            """,
            (template_id, invoice_number),
        ).fetchone()
    if not row:
        return None
    return history_record(row)


def list_users() -> list[dict]:
    with db_connect() as conn:
        rows = conn.execute(
            """
            SELECT id, username, display_name, role, active, created_at
            FROM users
            ORDER BY id
            """
        ).fetchall()
    return [
        {
            "id": row["id"],
            "username": row["username"],
            "displayName": row["display_name"],
            "role": row["role"],
            "active": bool(row["active"]),
            "createdAt": row["created_at"],
        }
        for row in rows
    ]


def get_user_by_id(user_id: int) -> dict | None:
    with db_connect() as conn:
        row = conn.execute(
            "SELECT id, username, display_name, role, active FROM users WHERE id = ?",
            (user_id,),
        ).fetchone()
    return public_user(row) if row else None


def update_user(admin: dict, user_id: int, payload: dict) -> tuple[dict | None, str | None]:
    with db_connect() as conn:
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if not row:
            return None, "Usuario no encontrado."

        display_name = clean(payload.get("displayName")) or row["display_name"]
        role = clean(payload.get("role")) or row["role"]
        if role not in VALID_ROLES:
            return None, "Rol no permitido."

        active = payload.get("active")
        active = 1 if active in (True, 1, "1", "true", "on") else 0
        if user_id == admin["id"] and not active:
            return None, "No puedes desactivar tu propio usuario."
        if user_id == admin["id"] and role != ROLE_SUPER_ADMIN:
            return None, "No puedes quitarte el rol de super admin."

        password = str(payload.get("password") or "")
        if password:
            if len(password) < 4:
                return None, "La clave debe tener minimo 4 caracteres."
            salt, password_hash = hash_password(password)
            conn.execute(
                """
                UPDATE users
                SET display_name = ?, role = ?, active = ?, password_salt = ?, password_hash = ?
                WHERE id = ?
                """,
                (display_name, role, active, salt, password_hash, user_id),
            )
        else:
            conn.execute(
                "UPDATE users SET display_name = ?, role = ?, active = ? WHERE id = ?",
                (display_name, role, active, user_id),
            )
        updated = conn.execute(
            "SELECT id, username, display_name, role, active FROM users WHERE id = ?",
            (user_id,),
        ).fetchone()
    return public_user(updated), None


def user_activity(user_id: int, limit: int = 25) -> dict:
    with db_connect() as conn:
        exports = conn.execute(
            """
            SELECT r.id, r.template_id, r.invoice_number, r.filename, r.row_count, r.created_at,
                   u.username, u.display_name
            FROM invoice_records r
            JOIN users u ON u.id = r.user_id
            WHERE r.user_id = ?
            ORDER BY r.id DESC
            LIMIT ?
            """,
            (user_id, limit),
        ).fetchall()
        drafts = conn.execute(
            """
            SELECT d.id, d.template_id, d.invoice_number, d.row_count, d.created_at, d.updated_at,
                   u.username, u.display_name
            FROM drafts d
            JOIN users u ON u.id = d.user_id
            WHERE d.user_id = ?
            ORDER BY d.updated_at DESC
            LIMIT ?
            """,
            (user_id, limit),
        ).fetchall()
        totals = conn.execute(
            """
            SELECT
              (SELECT COUNT(*) FROM invoice_records WHERE user_id = ?) AS exports_total,
              (SELECT COUNT(*) FROM drafts WHERE user_id = ?) AS drafts_total
            """,
            (user_id, user_id),
        ).fetchone()
    return {
        "exports": [history_record(row) for row in exports],
        "drafts": [draft_record(row) for row in drafts],
        "totals": {
            "exports": totals["exports_total"],
            "drafts": totals["drafts_total"],
        },
    }


def history_record(row: dict | sqlite3.Row) -> dict:
    return {
        "id": row["id"],
        "templateId": row["template_id"],
        "invoiceNumber": row["invoice_number"],
        "filename": row["filename"],
        "rowCount": row["row_count"],
        "createdAt": row["created_at"],
        "username": row["username"],
        "displayName": row["display_name"],
    }


def list_history(filters: dict | None = None, limit: int = 200) -> list[dict]:
    filters = filters or {}
    clauses = []
    params: list[str] = []

    query = clean(filters.get("q"))
    if query:
        like = f"%{query}%"
        op = like_op()
        clauses.append(f"(r.invoice_number {op} ? OR r.filename {op} ? OR u.username {op} ? OR u.display_name {op} ?)")
        params.extend([like, like, like, like])

    template_id = clean(filters.get("templateId"))
    if template_id in SCHEMA["templates"]:
        clauses.append("r.template_id = ?")
        params.append(template_id)

    username = clean(filters.get("username"))
    if username:
        clauses.append(f"u.username {like_op()} ?")
        params.append(f"%{username}%")

    date_from = clean(filters.get("dateFrom"))
    if date_from:
        clauses.append("r.created_at >= ?")
        params.append(f"{date_from}T00:00:00")

    date_to = clean(filters.get("dateTo"))
    if date_to:
        clauses.append("r.created_at <= ?")
        params.append(f"{date_to}T23:59:59")

    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    with db_connect() as conn:
        rows = conn.execute(
            f"""
            SELECT r.id, r.template_id, r.invoice_number, r.filename, r.row_count, r.created_at,
                   u.username, u.display_name
            FROM invoice_records r
            JOIN users u ON u.id = r.user_id
            {where}
            ORDER BY r.id DESC
            LIMIT ?
            """,
            (*params, limit),
        ).fetchall()
    return [history_record(row) for row in rows]


def draft_record(row: dict | sqlite3.Row) -> dict:
    return {
        "id": row["id"],
        "templateId": row["template_id"],
        "invoiceNumber": row["invoice_number"],
        "rowCount": row["row_count"],
        "createdAt": row["created_at"],
        "updatedAt": row["updated_at"],
        "username": row["username"],
        "displayName": row["display_name"],
    }


def list_drafts(user: dict, filters: dict | None = None, limit: int = 200) -> list[dict]:
    filters = filters or {}
    clauses = ["d.user_id = ?"]
    params: list = [user["id"]]

    query = clean(filters.get("q"))
    if query:
        like = f"%{query}%"
        op = like_op()
        clauses.append(f"(d.invoice_number {op} ? OR d.template_id {op} ?)")
        params.extend([like, like])

    template_id = clean(filters.get("templateId"))
    if template_id in SCHEMA["templates"]:
        clauses.append("d.template_id = ?")
        params.append(template_id)

    where = " AND ".join(clauses)
    with db_connect() as conn:
        rows = conn.execute(
            f"""
            SELECT d.id, d.template_id, d.invoice_number, d.row_count, d.created_at, d.updated_at,
                   u.username, u.display_name
            FROM drafts d
            JOIN users u ON u.id = d.user_id
            WHERE {where}
            ORDER BY d.updated_at DESC, d.id DESC
            LIMIT ?
            """,
            (*params, limit),
        ).fetchall()
    return [draft_record(row) for row in rows]


def get_draft(user: dict, draft_id: int) -> dict | None:
    with db_connect() as conn:
        row = conn.execute(
            """
            SELECT d.*, u.username, u.display_name
            FROM drafts d
            JOIN users u ON u.id = d.user_id
            WHERE d.id = ? AND d.user_id = ?
            """,
            (draft_id, user["id"]),
        ).fetchone()
    if not row:
        return None
    record = draft_record(row)
    record["payload"] = json.loads(row["payload_json"])
    return record


def save_draft(user: dict, template_id: str, rows: list[dict], draft_id: int | None = None) -> dict:
    payload = {"row": rows[0]} if SCHEMA["templates"][template_id]["mode"] == "single" else {"rows": rows}
    payload_json = json.dumps(payload, ensure_ascii=False)
    invoice_number = rows[0].get("NUM_FACTURA", "")
    timestamp = now_iso()
    with db_connect() as conn:
        if draft_id:
            row = conn.execute(
                "SELECT id FROM drafts WHERE id = ? AND user_id = ?",
                (draft_id, user["id"]),
            ).fetchone()
            if row:
                conn.execute(
                    """
                    UPDATE drafts
                    SET template_id = ?, invoice_number = ?, payload_json = ?, row_count = ?, updated_at = ?
                    WHERE id = ? AND user_id = ?
                    """,
                    (template_id, invoice_number, payload_json, len(rows), timestamp, draft_id, user["id"]),
                )
            else:
                draft_id = None
        if not draft_id:
            existing = None
            if invoice_number:
                existing = conn.execute(
                    """
                    SELECT id FROM drafts
                    WHERE user_id = ? AND template_id = ? AND invoice_number = ?
                    """,
                    (user["id"], template_id, invoice_number),
                ).fetchone()
            if existing:
                draft_id = existing["id"]
                conn.execute(
                    """
                    UPDATE drafts
                    SET payload_json = ?, row_count = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (payload_json, len(rows), timestamp, draft_id),
                )
            else:
                draft_id = conn.insert_and_get_id(
                    """
                    INSERT INTO drafts (user_id, template_id, invoice_number, payload_json, row_count, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (user["id"], template_id, invoice_number, payload_json, len(rows), timestamp, timestamp),
                )
        row = conn.execute(
            """
            SELECT d.id, d.template_id, d.invoice_number, d.row_count, d.created_at, d.updated_at,
                   u.username, u.display_name
            FROM drafts d
            JOIN users u ON u.id = d.user_id
            WHERE d.id = ?
            """,
            (draft_id,),
        ).fetchone()
    return draft_record(row)


def delete_draft(user: dict, draft_id: int) -> bool:
    with db_connect() as conn:
        cursor = conn.execute(
            "DELETE FROM drafts WHERE id = ? AND user_id = ?",
            (draft_id, user["id"]),
        )
    return cursor.rowcount > 0


def record_frequent_values(user: dict, rows: list[dict]):
    timestamp = now_iso()
    use_count_increment = "frequent_values.use_count + 1" if USE_POSTGRES else "use_count + 1"
    with db_connect() as conn:
        for row in rows:
            for field_name in FREQUENT_FIELD_NAMES:
                value = clean(row.get(field_name))
                if not value:
                    continue
                conn.execute(
                    f"""
                    INSERT INTO frequent_values (user_id, field_name, value, label, use_count, last_used_at)
                    VALUES (?, ?, ?, ?, 1, ?)
                    ON CONFLICT(user_id, field_name, value)
                    DO UPDATE SET
                        use_count = {use_count_increment},
                        label = excluded.label,
                        last_used_at = excluded.last_used_at
                    """,
                    (user["id"], field_name, value, value, timestamp),
                )


def list_frequent_values(user: dict, field_name: str, query: str = "", limit: int = 12) -> list[dict]:
    field_name = clean(field_name)
    if field_name not in FREQUENT_FIELD_NAMES:
        return []
    params: list = [user["id"], field_name]
    where = "user_id = ? AND field_name = ?"
    query = clean(query)
    if query:
        where += f" AND value {like_op()} ?"
        params.append(f"%{query}%")
    with db_connect() as conn:
        rows = conn.execute(
            f"""
            SELECT field_name, value, label, use_count, last_used_at
            FROM frequent_values
            WHERE {where}
            ORDER BY use_count DESC, last_used_at DESC
            LIMIT ?
            """,
            (*params, limit),
        ).fetchall()
    return [
        {
            "fieldName": row["field_name"],
            "value": row["value"],
            "label": row["label"],
            "useCount": row["use_count"],
            "lastUsedAt": row["last_used_at"],
        }
        for row in rows
    ]


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_invoice_number(value: str) -> str:
    value = clean(value).upper().replace(" ", "")
    if not value:
        return ""
    if value.startswith(INVOICE_PREFIX):
        return value
    return f"{INVOICE_PREFIX}{value}"


def parse_date(value: str):
    if not DATE_RE.match(value):
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def as_int(value: str):
    if not value or not DIGITS_RE.match(value):
        return None
    try:
        return int(value)
    except ValueError:
        return None


def get_fields(template_id: str) -> list[dict]:
    template = SCHEMA["templates"].get(template_id)
    if not template:
        raise KeyError(template_id)
    return template["fields"]


def condition(name: str | None, row: dict, template_id: str) -> bool:
    if not name:
        return False

    nature = clean(row.get("Naturaleza_del_evento"))
    state = clean(row.get("Estado_de_aseguramiento"))
    attention = clean(row.get("Es_atencion_inicial_paciente_remitido_o_control"))
    owner_doc = clean(row.get("Tipo_de_documento_de_identidad_del_propietario"))
    victim_doc = clean(row.get("Tipo_documento_identidad_victima"))
    service_type = clean(row.get("Tipo_de_servicio"))
    accident = nature == "01"
    policy_required = accident and state in {"4", "6"}
    owner_required = accident and state in {"2", "4", "6", "8"}
    owner_person_required = owner_required and owner_doc != "NI"
    driver_required = accident and state in {"2", "4", "6", "7", "8"}
    secondary = attention in {"3", "7", "8"}
    primary = attention in {"2", "6", "8"}

    event_date = parse_date(clean(row.get("Fecha_de_ocurrencia_evento")))
    siras_required = bool(
        accident and event_date and event_date > datetime(2023, 6, 1).date()
    )

    checks = {
        "accident": accident,
        "notAccident": not accident,
        "natureOther": nature == "17",
        "notNatureOther": nature != "17",
        "poblacionEspecialRequerida": victim_doc in {"AS", "MS"},
        "plateRequired": accident and state in {"2", "4", "6", "7"},
        "plateBlank": (not accident) or state in {"3", "8"},
        "policyRequired": policy_required,
        "notPolicyRequired": not policy_required,
        "sirasRequired": siras_required,
        "topRequired": accident and state == "6",
        "ownerRequired": owner_required,
        "notOwnerRequired": not owner_required,
        "ownerPersonRequired": owner_person_required,
        "ownerAddressBlank": (not owner_required) or owner_doc == "NI",
        "driverRequired": driver_required,
        "notDriverRequired": not driver_required,
        "secondaryTransport": secondary,
        "notSecondaryTransport": not secondary,
        "primaryTransport": primary,
        "notPrimaryTransport": not primary,
        "procedureService": service_type == "2",
        "nonProcedureService": service_type != "2",
        "serviceCodeRequired": service_type in {"1", "2", "5", "6", "7"},
        "serviceCodeVisible": service_type not in {"3", "4", "8"},
        "serviceCodeBlank": service_type in {"3", "4", "8"},
        "cupsVisible": service_type not in {"1", "5", "6", "7"},
        "cupsBlank": service_type in {"1", "5", "6", "7"},
    }
    return checks.get(name, False)


def is_required(field: dict, row: dict, template_id: str) -> bool:
    return bool(field.get("required")) or condition(field.get("requiredWhen"), row, template_id)


def should_empty(field: dict, row: dict, template_id: str) -> bool:
    return condition(field.get("emptyWhen"), row, template_id)


def normalize_row(template_id: str, input_row: dict) -> dict:
    fields = get_fields(template_id)
    row = {field["name"]: clean(input_row.get(field["name"])) for field in fields}
    if "NUM_FACTURA" in row:
        row["NUM_FACTURA"] = normalize_invoice_number(row["NUM_FACTURA"])

    for field in fields:
        if should_empty(field, row, template_id):
            row[field["name"]] = ""

    if template_id == "fur":
        if condition("accident", row, template_id) and clean(row.get("Estado_de_aseguramiento")):
            if clean(row.get("Estado_de_aseguramiento")) != "6":
                row["Cobro_por_agotamiento_tope_Aseguradora"] = "0"
        if condition("notAccident", row, template_id):
            row["Cobro_por_agotamiento_tope_Aseguradora"] = ""

    if template_id == "ser":
        qty = as_int(row.get("Cantidad_de_servicios", ""))
        billed = as_int(row.get("Valor_unitario_facturado", ""))
        claimed = as_int(row.get("Valor_unitario_reclamado", ""))
        if qty is not None and billed is not None:
            row["Valor_total_facturado"] = str(qty * billed)
        if qty is not None and claimed is not None:
            row["Valor_total_reclamado"] = str(qty * claimed)

    return row


def validate_field(field: dict, value: str, row: dict, template_id: str, row_number: int) -> list[dict]:
    errors = []
    name = field["name"]
    label = field.get("label", name)

    def add(message: str):
        errors.append({"row": row_number, "field": name, "label": label, "message": message})

    if is_required(field, row, template_id) and not value:
        add("Campo obligatorio segun las reglas seleccionadas.")
        return errors
    if not value:
        return errors

    max_length = field.get("maxLength")
    min_length = field.get("minLength")
    if max_length and len(value) > max_length:
        add(f"Maximo {max_length} caracteres.")
    if min_length and len(value) < min_length:
        add(f"Minimo {min_length} caracteres.")

    field_type = field.get("type")
    if field_type in {"numeric", "amount"} and not DIGITS_RE.match(value):
        add("Solo se permiten digitos, sin puntos, comas ni espacios.")
    if name in DIVIPOLA_FIELD_NAMES and DIVIPOLA_CODES and value not in DIVIPOLA_CODES:
        add("Seleccione un codigo DIVIPOLA valido.")
    if field_type == "amount":
        amount = as_int(value)
        min_value = field.get("minValue")
        if amount is not None and min_value is not None and amount < int(min_value):
            add(f"Debe ser mayor o igual a {min_value}.")
    if field_type == "date" and not parse_date(value):
        add("Use una fecha valida en formato AAAA-MM-DD.")
    if field_type == "time" and not TIME_RE.match(value):
        add("Use hora valida en formato HH:MM de 24 horas.")
    if field_type == "select":
        allowed = {item["value"] for item in SCHEMA["options"][field["optionsRef"]]}
        if value not in allowed:
            add("Seleccione un valor permitido.")
    if field.get("noComma") and "," in value:
        add("La guia indica que este campo no debe incluir coma (,).")

    return errors


def validate_cross_rules(template_id: str, row: dict, row_number: int) -> list[dict]:
    errors = []

    def add(field: str, message: str):
        label = next((f.get("label", field) for f in get_fields(template_id) if f["name"] == field), field)
        errors.append({"row": row_number, "field": field, "label": label, "message": message})

    if template_id == "fur":
        start = parse_date(clean(row.get("Fecha_de_inicio_de_vigencia_de_la_poliza")))
        end = parse_date(clean(row.get("Fecha_final_de_vigencia_de_la_poliza")))
        if start and end and end < start:
            add("Fecha_final_de_vigencia_de_la_poliza", "La fecha final de la poliza no puede ser anterior al inicio.")
        if condition("accident", row, template_id) and clean(row.get("Estado_de_aseguramiento")) in {"2", "3", "4", "7", "8"}:
            if clean(row.get("Cobro_por_agotamiento_tope_Aseguradora")) not in {"", "0"}:
                add("Cobro_por_agotamiento_tope_Aseguradora", "Para este estado de aseguramiento debe ser 0.")

    if template_id == "ser":
        unit_billed = as_int(row.get("Valor_unitario_facturado", ""))
        unit_claimed = as_int(row.get("Valor_unitario_reclamado", ""))
        total_billed = as_int(row.get("Valor_total_facturado", ""))
        total_claimed = as_int(row.get("Valor_total_reclamado", ""))
        if unit_billed is not None and unit_claimed is not None and unit_claimed > unit_billed:
            add("Valor_unitario_reclamado", "No puede ser mayor al valor unitario facturado.")
        if total_billed is not None and total_claimed is not None and total_claimed > total_billed:
            add("Valor_total_reclamado", "No puede ser mayor al valor total facturado.")

    return errors


def rows_from_payload(template_id: str, payload: dict) -> list[dict]:
    mode = SCHEMA["templates"][template_id]["mode"]
    if mode == "single":
        row = payload.get("row", payload)
        return [row] if isinstance(row, dict) else [{}]
    rows = payload.get("rows", [])
    return rows if isinstance(rows, list) and rows else [{}]


def validate_payload(template_id: str, payload: dict) -> tuple[list[dict], list[dict]]:
    if template_id not in SCHEMA["templates"]:
        return [], [{"row": 1, "field": "", "label": "Formulario", "message": "Formulario no reconocido."}]

    normalized_rows = []
    all_errors = []
    for index, raw_row in enumerate(rows_from_payload(template_id, payload), start=1):
        row = normalize_row(template_id, raw_row if isinstance(raw_row, dict) else {})
        normalized_rows.append(row)
        for field in get_fields(template_id):
            all_errors.extend(validate_field(field, row[field["name"]], row, template_id, index))
        all_errors.extend(validate_cross_rules(template_id, row, index))
    return normalized_rows, all_errors


def normalize_payload_rows(template_id: str, payload: dict) -> list[dict]:
    if template_id not in SCHEMA["templates"]:
        raise KeyError(template_id)
    normalized_rows = []
    for raw_row in rows_from_payload(template_id, payload):
        normalized_rows.append(normalize_row(template_id, raw_row if isinstance(raw_row, dict) else {}))
    return normalized_rows


def export_workbook(template_id: str, rows: list[dict]) -> tuple[bytes, str]:
    template = SCHEMA["templates"][template_id]
    path = TEMPLATE_DIR / template["file"]
    workbook = get_openpyxl().load_workbook(path)
    sheet = workbook.active

    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    expected = [field["name"] for field in get_fields(template_id)]
    if headers != expected:
        raise ValueError("Los encabezados de la plantilla no coinciden con el esquema.")

    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)

    for row_index, row in enumerate(rows, start=2):
        for col_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row_index, column=col_index)
            cell.value = clean(row.get(header))
            cell.number_format = "@"

    filename = template["file"]

    if SAVE_EXPORT_COPY_ENABLED:
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        local_copy = EXPORT_DIR / filename
        workbook.save(local_copy)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue(), filename


class AppHandler(BaseHTTPRequestHandler):
    server_version = "AdresAssistant/0.1"

    def log_message(self, fmt, *args):
        print("[%s] %s" % (self.log_date_time_string(), fmt % args))

    def handle_one_request(self):
        try:
            super().handle_one_request()
        except Exception as exc:
            self.send_server_error(exc)

    def send_server_error(self, exc: Exception):
        print(f"Error en la peticion: {exc.__class__.__name__}: {safe_error_text(exc)}")
        self.send_json(
            {
                "ok": False,
                "errors": [
                    {
                        "message": (
                            "La funcion inicio, pero hubo un error interno. "
                            "Revisa DATABASE_URL en Vercel y los logs del despliegue."
                        ),
                        "detail": f"{exc.__class__.__name__}: {safe_error_text(exc)}",
                    }
                ],
            },
            status=500,
        )

    def send_json(self, payload: dict, status: int = 200):
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def set_session_cookie(self, session_id: str):
        secure = "; Secure" if IS_VERCEL else ""
        self.send_header("Set-Cookie", f"{SESSION_COOKIE}={session_id}; Path=/; SameSite=Lax; HttpOnly{secure}")

    def clear_session_cookie(self):
        secure = "; Secure" if IS_VERCEL else ""
        self.send_header("Set-Cookie", f"{SESSION_COOKIE}=; Path=/; Max-Age=0; SameSite=Lax; HttpOnly{secure}")

    def send_auth_json(self, payload: dict, session_id: str | None = None, clear: bool = False, status: int = 200):
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        if clear:
            self.clear_session_cookie()
        elif session_id:
            self.set_session_cookie(session_id)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_static(self, path: Path):
        if not path.exists() or not path.is_file():
            self.send_error(404)
            return
        data = path.read_bytes()
        content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def read_json(self) -> dict:
        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length) if length else b"{}"
        if not raw:
            return {}
        return json.loads(raw.decode("utf-8"))

    def current_user(self) -> dict | None:
        cookie = SimpleCookie(self.headers.get("Cookie", ""))
        morsel = cookie.get(SESSION_COOKIE)
        if not morsel:
            return None
        return get_user_by_session(morsel.value)

    def require_user(self) -> dict | None:
        user = self.current_user()
        if not user:
            self.send_json({"ok": False, "errors": [{"message": "Inicia sesion para continuar."}]}, status=401)
            return None
        return user

    def require_super_admin(self) -> dict | None:
        user = self.require_user()
        if not user:
            return None
        if not user.get("isSuperAdmin"):
            self.send_json({"ok": False, "errors": [{"message": "Solo el super admin puede hacer esta accion."}]}, status=403)
            return None
        return user

    def do_GET(self):
        try:
            self.route_GET()
        except Exception as exc:
            self.send_server_error(exc)

    def route_GET(self):
        parsed = urlparse(self.path)
        path = unquote(parsed.path)
        query = {key: values[0] for key, values in parse_qs(parsed.query).items() if values}
        if path == "/":
            self.send_static(STATIC_DIR / "index.html")
            return
        if path == "/api/schema":
            self.send_json(deepcopy(SCHEMA))
            return
        if path == "/api/divipola":
            self.send_json({"items": DIVIPOLA_ITEMS})
            return
        if path == "/api/health":
            ensure_db()
            self.send_json(
                {
                    "ok": True,
                    "database": "postgres" if USE_POSTGRES else "sqlite",
                    "databaseUrlConfigured": bool(DATABASE_URL),
                }
            )
            return
        if path == "/api/session":
            self.send_json({"user": self.current_user(), "setupRequired": user_count() == 0})
            return
        if path == "/api/history":
            if not self.require_user():
                return
            self.send_json({"records": list_history(query)})
            return
        if path == "/api/drafts":
            user = self.require_user()
            if not user:
                return
            self.send_json({"drafts": list_drafts(user, query)})
            return
        if path == "/api/frequent":
            user = self.require_user()
            if not user:
                return
            self.send_json(
                {
                    "values": list_frequent_values(
                        user,
                        query.get("fieldName", ""),
                        query.get("q", ""),
                    )
                }
            )
            return
        if path.startswith("/api/drafts/"):
            user = self.require_user()
            if not user:
                return
            try:
                draft_id = int(path.removeprefix("/api/drafts/"))
            except ValueError:
                self.send_json({"ok": False, "errors": [{"message": "Borrador no reconocido."}]}, status=404)
                return
            draft = get_draft(user, draft_id)
            if not draft:
                self.send_json({"ok": False, "errors": [{"message": "Borrador no encontrado."}]}, status=404)
                return
            self.send_json({"draft": draft})
            return
        if path == "/api/users":
            if not self.require_super_admin():
                return
            self.send_json({"users": list_users()})
            return
        if path.startswith("/api/users/") and path.endswith("/activity"):
            if not self.require_super_admin():
                return
            try:
                user_id = int(path.removeprefix("/api/users/").removesuffix("/activity"))
            except ValueError:
                self.send_json({"ok": False, "errors": [{"message": "Usuario no reconocido."}]}, status=404)
                return
            self.send_json({"activity": user_activity(user_id)})
            return
        if path.startswith("/static/"):
            target = (STATIC_DIR / path.removeprefix("/static/")).resolve()
            if not str(target).startswith(str(STATIC_DIR.resolve())):
                self.send_error(403)
                return
            self.send_static(target)
            return
        self.send_error(404)

    def do_POST(self):
        try:
            self.route_POST()
        except Exception as exc:
            self.send_server_error(exc)

    def route_POST(self):
        path = unquote(urlparse(self.path).path)
        if path in {"/api/login", "/api/register", "/api/logout"}:
            self.handle_auth(path)
            return

        if path.startswith("/api/drafts/"):
            self.handle_save_draft(path)
            return
        if path.startswith("/api/users/"):
            self.handle_update_user(path)
            return

        parts = [part for part in path.split("/") if part]
        if len(parts) != 3 or parts[0] != "api" or parts[1] not in {"validate", "export"}:
            self.send_error(404)
            return

        template_id = parts[2]
        try:
            payload = self.read_json()
            rows, errors = validate_payload(template_id, payload)
        except json.JSONDecodeError:
            self.send_json({"errors": [{"message": "JSON invalido."}]}, status=400)
            return
        except KeyError:
            self.send_json({"errors": [{"message": "Formulario no reconocido."}]}, status=404)
            return

        if parts[1] == "validate":
            self.send_json({"ok": not errors, "rows": rows, "errors": errors})
            return

        user = self.require_user()
        if not user:
            return

        if errors:
            self.send_json({"ok": False, "errors": errors}, status=422)
            return

        try:
            invoice_number = rows[0].get("NUM_FACTURA", "")
            data, filename = export_workbook(template_id, rows)
            record_export(user, template_id, invoice_number, filename, len(rows))
            record_frequent_values(user, rows)
        except Exception as exc:
            self.send_json({"ok": False, "errors": [{"message": str(exc)}]}, status=500)
            return

        self.send_response(200)
        self.send_header(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_DELETE(self):
        try:
            self.route_DELETE()
        except Exception as exc:
            self.send_server_error(exc)

    def route_DELETE(self):
        path = unquote(urlparse(self.path).path)
        if not path.startswith("/api/drafts/"):
            self.send_error(404)
            return
        user = self.require_user()
        if not user:
            return
        try:
            draft_id = int(path.removeprefix("/api/drafts/"))
        except ValueError:
            self.send_json({"ok": False, "errors": [{"message": "Borrador no reconocido."}]}, status=404)
            return
        if not delete_draft(user, draft_id):
            self.send_json({"ok": False, "errors": [{"message": "Borrador no encontrado."}]}, status=404)
            return
        self.send_json({"ok": True})

    def handle_save_draft(self, path: str):
        user = self.require_user()
        if not user:
            return
        parts = [part for part in path.split("/") if part]
        if len(parts) != 3 or parts[0] != "api" or parts[1] != "drafts":
            self.send_error(404)
            return
        template_id = parts[2]
        try:
            payload = self.read_json()
            rows = normalize_payload_rows(template_id, payload)
        except json.JSONDecodeError:
            self.send_json({"ok": False, "errors": [{"message": "JSON invalido."}]}, status=400)
            return
        except KeyError:
            self.send_json({"ok": False, "errors": [{"message": "Formulario no reconocido."}]}, status=404)
            return
        draft_id = payload.get("draftId")
        try:
            draft_id = int(draft_id) if draft_id else None
        except (TypeError, ValueError):
            draft_id = None
        draft = save_draft(user, template_id, rows, draft_id)
        record_frequent_values(user, rows)
        self.send_json({"ok": True, "draft": draft})

    def handle_update_user(self, path: str):
        admin = self.require_super_admin()
        if not admin:
            return
        try:
            user_id = int(path.removeprefix("/api/users/"))
        except ValueError:
            self.send_json({"ok": False, "errors": [{"message": "Usuario no reconocido."}]}, status=404)
            return
        try:
            payload = self.read_json()
        except json.JSONDecodeError:
            self.send_json({"ok": False, "errors": [{"message": "JSON invalido."}]}, status=400)
            return
        user, error = update_user(admin, user_id, payload)
        if error:
            self.send_json({"ok": False, "errors": [{"message": error}]}, status=422)
            return
        self.send_json({"ok": True, "user": user})

    def handle_auth(self, path: str):
        if path == "/api/logout":
            cookie = SimpleCookie(self.headers.get("Cookie", ""))
            morsel = cookie.get(SESSION_COOKIE)
            if morsel:
                delete_session(morsel.value)
            self.send_auth_json({"ok": True, "user": None}, clear=True)
            return

        try:
            payload = self.read_json()
        except json.JSONDecodeError:
            self.send_json({"ok": False, "errors": [{"message": "JSON invalido."}]}, status=400)
            return

        username = clean(payload.get("username"))
        password = str(payload.get("password") or "")
        if path == "/api/register":
            bootstrap = user_count() == 0
            role = clean(payload.get("role")) or ROLE_FACTURADOR
            if bootstrap:
                role = ROLE_SUPER_ADMIN
            else:
                admin = self.require_super_admin()
                if not admin:
                    return
            user, error = create_user(username, password, clean(payload.get("displayName")), role)
            if error:
                self.send_json({"ok": False, "errors": [{"message": error}]}, status=422)
                return
            if not bootstrap:
                self.send_json({"ok": True, "user": user})
                return
        else:
            user = authenticate(username, password)
            if not user:
                self.send_json({"ok": False, "errors": [{"message": "Usuario o clave incorrectos."}]}, status=401)
                return

        session_id = create_session(user)
        self.send_auth_json({"ok": True, "user": user}, session_id=session_id)


class WsgiAppHandler(AppHandler):
    def __init__(self, environ: dict, body: bytes):
        self.environ = environ
        self.command = environ.get("REQUEST_METHOD", "GET").upper()
        path = environ.get("PATH_INFO") or "/"
        query = environ.get("QUERY_STRING") or ""
        self.path = f"{path}?{query}" if query else path
        self.request_version = environ.get("SERVER_PROTOCOL", "HTTP/1.1")
        self.client_address = (environ.get("REMOTE_ADDR", ""), 0)
        self.server = None
        self.rfile = BytesIO(body)
        self.wfile = BytesIO()
        self.headers = self.make_headers(environ)
        self.status_code = 200
        self.response_headers: list[tuple[str, str]] = []

    @staticmethod
    def make_headers(environ: dict) -> Message:
        headers = Message()
        for key, value in environ.items():
            if key.startswith("HTTP_"):
                name = key.removeprefix("HTTP_").replace("_", "-").title()
                headers[name] = value
        if environ.get("CONTENT_TYPE"):
            headers["Content-Type"] = environ["CONTENT_TYPE"]
        if environ.get("CONTENT_LENGTH"):
            headers["Content-Length"] = environ["CONTENT_LENGTH"]
        return headers

    def send_response(self, code, message=None):
        self.status_code = int(code)

    def send_header(self, keyword, value):
        self.response_headers.append((keyword, str(value)))

    def end_headers(self):
        return

    def send_error(self, code, message=None, explain=None):
        self.send_json(
            {"ok": False, "errors": [{"message": message or HTTPStatus(code).phrase}]},
            status=code,
        )

    def run(self):
        try:
            if self.command == "GET":
                self.route_GET()
            elif self.command == "POST":
                self.route_POST()
            elif self.command == "DELETE":
                self.route_DELETE()
            else:
                self.send_error(405, "Metodo no permitido.")
        except Exception as exc:
            self.send_server_error(exc)
        return self.status_code, self.response_headers, self.wfile.getvalue()


def application(environ, start_response):
    try:
        length = int(environ.get("CONTENT_LENGTH") or 0)
    except ValueError:
        length = 0
    body = environ["wsgi.input"].read(length) if length else b""
    handler = WsgiAppHandler(environ, body)
    status_code, headers, data = handler.run()
    phrase = HTTPStatus(status_code).phrase if status_code in HTTPStatus._value2member_map_ else "OK"
    start_response(f"{status_code} {phrase}", headers)
    return [data]


app = application
handler = application


def main():
    parser = argparse.ArgumentParser(description="Asistente local para plantillas ADRES FUR.")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", default=8787, type=int)
    args = parser.parse_args()

    address = (args.host, args.port)
    httpd = ThreadingHTTPServer(address, AppHandler)
    print(f"Servidor listo en http://{args.host}:{args.port}")
    httpd.serve_forever()


if __name__ == "__main__":
    main()
